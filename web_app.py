#!/usr/bin/env python
"""
Vehicle License Plate Recognition Web Application
Features:
- User login/registration
- Vehicle number management (Excel integration)
- Plate recognition from uploaded images
- Email alerts when vehicle is detected
"""

import os
import json
import hashlib
import smtplib
from datetime import datetime
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from PIL import Image, ImageDraw, ImageFont

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook, load_workbook

# Lazy import for YOLO - defer to avoid PyTorch initialization issues
YOLO_AVAILABLE = False
YOLO = None

# Import plate recognition functions
try:
    from plate_recognition import (
        DEFAULT_API_KEY,
        process_full_image,
        recognition_api,
    )
except ImportError:
    print("Warning: Could not import plate_recognition. Some features may not work.")
    DEFAULT_API_KEY = "b21fc6a678b12c50da9aad4a72cde7fd879fb3b7"

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-in-production-2024')

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
EXCEL_FILE = 'vehicles.xlsx'
USERS_DB = 'users.db'
YOLO_WEIGHTS_PATH = os.getenv('YOLO_WEIGHTS_PATH', 'best.pt')

# Additional paths
ANNOTATED_FOLDER = os.path.join(UPLOAD_FOLDER, 'verification', 'annotated')

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'verification'), exist_ok=True)
os.makedirs(ANNOTATED_FOLDER, exist_ok=True)
os.makedirs('instance', exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ANNOTATED_FOLDER'] = ANNOTATED_FOLDER

# Email configuration (update in config.py)
try:
    from config import MAIL_SERVER, MAIL_PORT, MAIL_USE_TLS, MAIL_USERNAME, MAIL_PASSWORD
except ImportError:
    MAIL_SERVER = 'smtp.gmail.com'
    MAIL_PORT = 587
    MAIL_USE_TLS = True
    MAIL_USERNAME = 'your-email@gmail.com'  # Update this
    MAIL_PASSWORD = 'your-app-password'  # Update this


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def hash_password(password):
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()


def init_database():
    """Initialize SQLite database for users and verification logs"""
    import sqlite3
    conn = sqlite3.connect(USERS_DB)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  email TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS verification_logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  vehicle_number TEXT NOT NULL,
                  is_registered INTEGER DEFAULT 0,
                  image_path TEXT,
                  detected_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  user_email TEXT)''')
    conn.commit()
    conn.close()


def init_excel():
    """Initialize Excel file for vehicle numbers"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicles"
        ws.append(['Vehicle Number', 'Owner Name', 'Email', 'Registered Date', 'Status'])
        wb.save(EXCEL_FILE)
    return EXCEL_FILE


def read_vehicles_from_excel():
    """Read all vehicle numbers from Excel file"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        vehicles = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if row[0]:  # If vehicle number exists
                vehicles.append({
                    'number': str(row[0]).upper().strip(),
                    'owner': row[1] if row[1] else '',
                    'email': row[2] if row[2] else '',
                    'date': row[3] if row[3] else '',
                    'status': row[4] if row[4] else 'Active'
                })
        return vehicles
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return []


def add_vehicle_to_excel(vehicle_number, owner_name, email):
    """Add a new vehicle to Excel file"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            vehicle_number.upper().strip(),
            owner_name,
            email,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Active'
        ])
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error adding to Excel: {e}")
        return False


def check_vehicle_in_excel(vehicle_number):
    """Check if vehicle number exists in Excel"""
    vehicles = read_vehicles_from_excel()
    vehicle_number = vehicle_number.upper().strip()
    for vehicle in vehicles:
        if vehicle['number'] == vehicle_number:
            return vehicle
    return None


_yolo_model = None


def load_yolo_model():
    """Lazy load YOLO model if available."""
    global _yolo_model
    if not YOLO_AVAILABLE:
        return None
    if _yolo_model is None:
        weights_path = Path(YOLO_WEIGHTS_PATH)
        if not weights_path.exists():
            print(f"YOLO weights file not found at {weights_path}. Skipping YOLO detection.")
            return None
        try:
            _yolo_model = YOLO(str(weights_path))
            print(f"✓ Successfully loaded YOLO model from {weights_path}")
        except AttributeError as exc:
            error_msg = str(exc)
            if 'SCDown' in error_msg or 'custom' in error_msg.lower():
                print(f"⚠️ YOLO model compatibility issue: The model appears to use a custom architecture.")
                print(f"   This model may require a specific ultralytics version or custom modules.")
                print(f"   The app will continue without YOLO detection, using Plate Recognizer API instead.")
            else:
                print(f"⚠️ Failed to load YOLO model: {exc}")
            _yolo_model = None
        except Exception as exc:
            print(f"⚠️ Failed to load YOLO model from {weights_path}: {exc}")
            print(f"   The app will continue without YOLO detection, using Plate Recognizer API instead.")
            _yolo_model = None
    return _yolo_model


def detect_plate_yolo(image_path, conf_threshold=0.25, model=None):
    """Detect license plate bounding boxes using YOLO."""
    if model is None:
        model = load_yolo_model()
    if model is None:
        return []
    try:
        results = model(image_path)
    except Exception as exc:
        print(f"Error running YOLO detection: {exc}")
        return []

    boxes = []
    for result in results:
        if not hasattr(result, "boxes") or result.boxes is None:
            continue
        for box in result.boxes:
            try:
                confidence = float(box.conf[0]) if hasattr(box, "conf") else 0.0
                if confidence < conf_threshold:
                    continue
                x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                boxes.append(
                    {
                        "x1": max(0, x1),
                        "y1": max(0, y1),
                        "x2": max(0, x2),
                        "y2": max(0, y2),
                        "confidence": round(confidence, 3),
                    }
                )
            except Exception as exc:
                print(f"Error parsing YOLO box: {exc}")
    return boxes


def send_email_alert(to_email, vehicle_number, image_path, is_registered=True):
    """Send email alert when vehicle is detected"""
    # Check if email is configured
    if MAIL_USERNAME == 'your-email@gmail.com' or MAIL_PASSWORD == 'your-app-password':
        error_msg = "Email not configured! Please update config.py with your email credentials."
        print(f"ERROR: {error_msg}")
        return False, error_msg
    
    try:
        msg = MIMEMultipart()
        msg['From'] = MAIL_USERNAME
        msg['To'] = to_email
        
        detection_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if is_registered:
            msg['Subject'] = f"✅ Registered Vehicle Detected - {vehicle_number}"
            status_color = "#28a745"
            status_icon = "✅"
            status_text = "REGISTERED VEHICLE"
            alert_message = "This vehicle is registered in the system and has been verified."
        else:
            msg['Subject'] = f"⚠️ UNREGISTERED Vehicle Alert - {vehicle_number}"
            status_color = "#dc3545"
            status_icon = "⚠️"
            status_text = "UNREGISTERED VEHICLE"
            alert_message = "<strong style='color: #dc3545;'>WARNING:</strong> This vehicle is NOT registered in the system. Please take appropriate action."

        # Enhanced email body
        body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0; }}
                .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
                .alert-box {{ background: {status_color}; color: white; padding: 15px; border-radius: 5px; margin: 20px 0; text-align: center; font-size: 18px; font-weight: bold; }}
                .info-box {{ background: white; padding: 15px; border-left: 4px solid {status_color}; margin: 15px 0; }}
                .footer {{ text-align: center; margin-top: 20px; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>{status_icon} Vehicle Detection Alert</h1>
                </div>
                <div class="content">
                    <div class="alert-box">
                        {status_text}
                    </div>
                    
                    <div class="info-box">
                        <p><strong>Vehicle Number:</strong> <span style="font-size: 20px; color: {status_color};">{vehicle_number}</span></p>
                        <p><strong>Detection Time:</strong> {detection_time}</p>
                        <p><strong>Status:</strong> {status_text}</p>
                    </div>
                    
                    <div class="info-box">
                        <p>{alert_message}</p>
                    </div>
                    
                    <p style="margin-top: 20px;">The detected vehicle image has been attached to this email for your records.</p>
                    
                    <div class="footer">
                        <p>This is an automated alert from Vehicle License Plate Recognition System</p>
                        <p>Detection Time: {detection_time}</p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        # Attach image
        if os.path.exists(image_path):
            with open(image_path, 'rb') as f:
                img_data = f.read()
            image = MIMEImage(img_data)
            image.add_header('Content-Disposition', f'attachment; filename=vehicle_{vehicle_number}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg')
            msg.attach(image)

        # Send email
        print(f"Attempting to send email to {to_email}...")
        print(f"Using SMTP: {MAIL_SERVER}:{MAIL_PORT}")
        print(f"From: {MAIL_USERNAME}")
        
        server = smtplib.SMTP(MAIL_SERVER, MAIL_PORT)
        server.starttls()
        server.login(MAIL_USERNAME, MAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        print(f"✅ Email sent successfully to {to_email}")
        return True, "Email sent successfully"
    except smtplib.SMTPAuthenticationError as e:
        error_msg = f"Email authentication failed. Check your email and app password in config.py. Error: {str(e)}"
        print(f"ERROR: {error_msg}")
        return False, error_msg
    except smtplib.SMTPException as e:
        error_msg = f"SMTP error: {str(e)}"
        print(f"ERROR: {error_msg}")
        return False, error_msg
    except Exception as e:
        error_msg = f"Error sending email: {str(e)}"
        print(f"ERROR: {error_msg}")
        return False, error_msg


def log_verification(vehicle_number, is_registered, image_path, user_email):
    """Log vehicle verification to database"""
    import sqlite3
    try:
        conn = sqlite3.connect(USERS_DB)
        c = conn.cursor()
        c.execute('''INSERT INTO verification_logs 
                     (vehicle_number, is_registered, image_path, user_email)
                     VALUES (?, ?, ?, ?)''',
                  (vehicle_number, 1 if is_registered else 0, image_path, user_email))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error logging verification: {e}")
        return False


def get_verification_stats():
    """Get verification statistics"""
    import sqlite3
    try:
        conn = sqlite3.connect(USERS_DB)
        c = conn.cursor()
        
        # Today's count
        c.execute('''SELECT COUNT(*) FROM verification_logs 
                     WHERE DATE(detected_time) = DATE('now')''')
        today_count = c.fetchone()[0]
        
        # Total count
        c.execute('SELECT COUNT(*) FROM verification_logs')
        total_count = c.fetchone()[0]
        
        # Registered vs Unregistered today
        c.execute('''SELECT COUNT(*) FROM verification_logs 
                     WHERE DATE(detected_time) = DATE('now') AND is_registered = 1''')
        registered_today = c.fetchone()[0]
        
        c.execute('''SELECT COUNT(*) FROM verification_logs 
                     WHERE DATE(detected_time) = DATE('now') AND is_registered = 0''')
        unregistered_today = c.fetchone()[0]
        
        # Recent verifications (last 24 hours)
        c.execute('''SELECT vehicle_number, is_registered, detected_time 
                     FROM verification_logs 
                     WHERE detected_time >= datetime('now', '-24 hours')
                     ORDER BY detected_time DESC LIMIT 20''')
        recent = c.fetchall()
        
        conn.close()
        
        return {
            'today_count': today_count,
            'total_count': total_count,
            'registered_today': registered_today,
            'unregistered_today': unregistered_today,
            'recent': recent
        }
    except Exception as e:
        print(f"Error getting stats: {e}")
        return {
            'today_count': 0,
            'total_count': 0,
            'registered_today': 0,
            'unregistered_today': 0,
            'recent': []
        }


def get_daily_history():
    """Get daily vehicle history grouped by date"""
    import sqlite3
    try:
        conn = sqlite3.connect(USERS_DB)
        c = conn.cursor()
        
        # Get last 7 days of data
        c.execute('''SELECT 
                     DATE(detected_time) as date,
                     COUNT(*) as count,
                     GROUP_CONCAT(vehicle_number || '|' || TIME(detected_time) || '|' || is_registered, '||') as vehicles
                     FROM verification_logs
                     WHERE detected_time >= datetime('now', '-7 days')
                     GROUP BY DATE(detected_time)
                     ORDER BY date DESC''')
        
        daily_data = []
        for row in c.fetchall():
            date, count, vehicles_str = row
            vehicles = []
            if vehicles_str:
                for v in vehicles_str.split('||'):
                    parts = v.split('|')
                    if len(parts) >= 3:
                        vehicles.append({
                            'number': parts[0],
                            'time': parts[1],
                            'is_registered': bool(int(parts[2]))
                        })
            
            daily_data.append({
                'date': date,
                'count': count,
                'vehicles': vehicles
            })
        
        conn.close()
        return daily_data
    except Exception as e:
        print(f"Error getting daily history: {e}")
        return []


def annotate_image_with_boxes(image_path, detections, output_path=None):
    """Create an annotated image with bounding boxes.

    Accepts detections from either the Plate Recognizer API (with `box` keys)
    or YOLO detections (with `x1`, `y1`, `x2`, `y2` keys).
    """
    try:
        source = Path(image_path)
        image = Image.open(source).convert("RGB")
        draw = ImageDraw.Draw(image)
        try:
            font = ImageFont.truetype("arial.ttf", 18)
        except (OSError, IOError):
            font = ImageFont.load_default()

        for detection in detections:
            if not isinstance(detection, dict):
                continue

            if "box" in detection and detection["box"]:
                box = detection["box"]
                xmin, ymin, xmax, ymax = box["xmin"], box["ymin"], box["xmax"], box["ymax"]
                plate_text = detection.get("plate") or "PLATE"
                confidence = detection.get("score")
                label = f"{plate_text} ({confidence:.2f})" if confidence is not None else plate_text
                rect_color = (255, 82, 82)
            elif all(key in detection for key in ("x1", "y1", "x2", "y2")):
                xmin, ymin, xmax, ymax = detection["x1"], detection["y1"], detection["x2"], detection["y2"]
                confidence = detection.get("confidence")
                label = f"YOLO {confidence:.2f}" if confidence is not None else "YOLO"
                rect_color = (0, 200, 0)
            else:
                continue

            draw.rectangle([(xmin, ymin), (xmax, ymax)], outline=rect_color, width=4)

            text_size = draw.textbbox((0, 0), label, font=font)
            text_width = text_size[2] - text_size[0]
            text_height = text_size[3] - text_size[1]
            text_bg = [(xmin, ymin - text_height - 6), (xmin + text_width + 12, ymin)]
            draw.rectangle(text_bg, fill=rect_color)
            draw.text((xmin + 6, ymin - text_height - 4), label, fill=(255, 255, 255), font=font)

        if output_path is None:
            annotated_name = f"{source.stem}_annotated{source.suffix}"
            annotated_path = Path(ANNOTATED_FOLDER) / annotated_name
        else:
            annotated_path = Path(output_path)
            annotated_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(annotated_path)
        uploads_root = Path(app.config['UPLOAD_FOLDER']).resolve()
        try:
            relative_path = str(annotated_path.resolve().relative_to(uploads_root))
        except Exception:
            relative_path = str(annotated_path)
        return relative_path.replace('\\', '/')
    except Exception as e:
        print(f"Error annotating image: {e}")
        return None


def process_plate_recognition(image_path):
    """Process image using plate recognition API"""
    try:
        # Create args object for plate recognition
        class Args:
            def __init__(self):
                self.api_key = DEFAULT_API_KEY
                self.regions = []
                self.sdk_url = None
                self.camera_id = None
                self.mmc = False
                self.show_boxes = False
                self.annotate_images = False
                self.crop_lp = None
                self.crop_vehicle = None
                self.engine_config = None
                self.split_image = False
                self.split_x = 0
                self.split_y = 0
                self.split_overlap = 10

        args = Args()
        result = process_full_image(Path(image_path), args, {})
        return result
    except Exception as e:
        print(f"Error in plate recognition: {e}")
        return None


# Initialize on startup
init_database()
init_excel()


@app.route('/')
def index():
    """Redirect to dashboard directly"""
    return redirect(url_for('dashboard'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login - bypassed for authentication-free mode"""
    # Set default session values
    if 'user_id' not in session:
        session['user_id'] = 1
        session['user_name'] = 'Guest User'
        session['user_email'] = 'guest@example.com'
    return redirect(url_for('dashboard'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration - bypassed for authentication-free mode"""
    # Set default session values
    if 'user_id' not in session:
        session['user_id'] = 1
        session['user_name'] = 'Guest User'
        session['user_email'] = 'guest@example.com'
    return redirect(url_for('dashboard'))


@app.route('/logout')
def logout():
    """User logout - not used in authentication-free mode"""
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
def dashboard():
    """Dashboard page"""
    # Ensure user session exists
    if 'user_id' not in session:
        session['user_id'] = 1
        session['user_name'] = 'Guest User'
        session['user_email'] = 'guest@example.com'
    
    # Get statistics
    vehicles = read_vehicles_from_excel()
    total_vehicles = len(vehicles)
    
    # Get verification statistics
    stats = get_verification_stats()
    
    return render_template('dashboard.html', 
                         user_name=session.get('user_name'),
                         total_vehicles=total_vehicles,
                         stats=stats)


@app.route('/register_vehicle', methods=['GET', 'POST'])
def register_vehicle():
    """Register a new vehicle number"""
    # Ensure user session exists
    if 'user_id' not in session:
        session['user_id'] = 1
        session['user_name'] = 'Guest User'
        session['user_email'] = 'guest@example.com'
    
    if request.method == 'POST':
        vehicle_number = request.form.get('vehicle_number', '').strip().upper()
        owner_name = request.form.get('owner_name', '').strip()
        
        if not vehicle_number:
            flash('Vehicle number is required', 'error')
            return render_template('register_vehicle.html')
        
        # Check if vehicle already exists
        existing = check_vehicle_in_excel(vehicle_number)
        if existing:
            flash(f'Vehicle {vehicle_number} is already registered', 'error')
        else:
            if add_vehicle_to_excel(vehicle_number, owner_name, session.get('user_email', '')):
                flash(f'Vehicle {vehicle_number} registered successfully!', 'success')
            else:
                flash('Error registering vehicle', 'error')
        
        return redirect(url_for('register_vehicle'))
    
    vehicles = read_vehicles_from_excel()
    return render_template('register_vehicle.html', vehicles=vehicles)


@app.route('/verify_vehicle', methods=['GET', 'POST'])
def verify_vehicle():
    """Upload image and verify vehicle"""
    # Ensure user session exists
    if 'user_id' not in session:
        session['user_id'] = 1
        session['user_name'] = 'Guest User'
        session['user_email'] = 'guest@example.com'

    yolo_warning = None
    yolo_boxes = []
    annotated_rel_path = None
    latest_image_filename = None
    
    if request.method == 'POST':
        if 'image' not in request.files:
            flash('No image file selected', 'error')
            return redirect(url_for('verify_vehicle'))
        
        file = request.files['image']
        if file.filename == '':
            flash('No image file selected', 'error')
            return redirect(url_for('verify_vehicle'))
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'verification', filename)
            file.save(filepath)
            latest_image_filename = filename

            # YOLO detection and annotation
            if YOLO_AVAILABLE:
                yolo_model = load_yolo_model()
                if yolo_model:
                    yolo_boxes = detect_plate_yolo(filepath, model=yolo_model)
                    if yolo_boxes:
                        annotated_filename = f"{Path(filename).stem}_yolo{Path(filename).suffix}"
                        annotated_full_path = os.path.join(app.config['ANNOTATED_FOLDER'], annotated_filename)
                        saved_path = annotate_image_with_boxes(filepath, yolo_boxes, annotated_full_path)
                        if saved_path:
                            annotated_rel_path = saved_path
                    else:
                        yolo_warning = "YOLO model did not detect any license plates in this image."
                else:
                    # Check if file exists but failed to load
                    weights_path = Path(YOLO_WEIGHTS_PATH)
                    if weights_path.exists():
                        yolo_warning = (
                            f"YOLO model file found at {YOLO_WEIGHTS_PATH} but could not be loaded. "
                            "This may be due to model compatibility issues. The app will use Plate Recognizer API instead."
                        )
                    else:
                        yolo_warning = (
                            f"YOLO weights not found at {YOLO_WEIGHTS_PATH}. "
                            "The app will use Plate Recognizer API instead."
                        )
            else:
                yolo_warning = "YOLO dependency not installed. Run 'pip install ultralytics opencv-python'."
            
            # Process image for plate recognition
            flash('Processing image... Please wait', 'info')
            result = process_plate_recognition(filepath)
            
            if result and 'results' in result and result['results']:
                detected_plates = []
                for plate_data in result['results']:
                    plate_text = plate_data.get('plate', '').strip().upper()
                    if plate_text:
                        detected_plates.append(plate_text)
                
                if detected_plates:
                    # Check each detected plate
                    alerts_sent = []
                    if not annotated_rel_path:
                        annotated_rel_path = annotate_image_with_boxes(filepath, result.get("results", []))
                    for plate_number in detected_plates:
                        vehicle_info = check_vehicle_in_excel(plate_number)
                        
                        if vehicle_info:
                            # Vehicle is registered - just log, NO EMAIL
                            email_to = vehicle_info.get('email') or session.get('user_email')
                            # Log the verification
                            log_verification(plate_number, True, filename, email_to)
                            alerts_sent.append(f"{plate_number} (Registered - No alert needed)")
                        else:
                            # Vehicle NOT registered - SEND EMAIL ALERT to logged-in user
                            email_to = session.get('user_email')
                            # Log the verification
                            log_verification(plate_number, False, filename, email_to)
                            # Send detailed email alert for UNREGISTERED vehicle
                            email_result, email_message = send_email_alert(email_to, plate_number, filepath, is_registered=False)
                            if email_result:
                                alerts_sent.append(f"{plate_number} (⚠️ UNREGISTERED - Alert sent to {email_to})")
                            else:
                                alerts_sent.append(f"{plate_number} (⚠️ UNREGISTERED - Email failed: {email_message})")
                                # Show error to user
                                flash(f'Email Error: {email_message}', 'error')
                    
                    flash(f'Detection complete! Plates found: {", ".join(detected_plates)}', 'success')
                    
                    # Show different messages for registered vs unregistered
                    unregistered = [a for a in alerts_sent if 'UNREGISTERED' in a]
                    registered = [a for a in alerts_sent if 'Registered' in a]
                    
                    if unregistered:
                        flash(f'⚠️ UNREGISTERED VEHICLES DETECTED! Email alerts sent: {", ".join(unregistered)}', 'danger')
                    if registered:
                        flash(f'✅ Registered vehicles: {", ".join(registered)}', 'success')
                    
                    return render_template(
                        'verify_vehicle.html',
                        result=result,
                        detected_plates=detected_plates,
                        image_path=latest_image_filename,
                        yolo_boxes=yolo_boxes,
                        annotated_image=annotated_rel_path,
                        yolo_available=YOLO_AVAILABLE,
                        yolo_warning=yolo_warning,
                        yolo_weights=YOLO_WEIGHTS_PATH,
                    )
                else:
                    flash('No license plates detected in the image', 'warning')
            else:
                flash('Could not detect any license plates. Please try another image.', 'warning')
        else:
            flash('Invalid file type. Please upload an image file (jpg, png, etc.)', 'error')
    
    return render_template(
        'verify_vehicle.html',
        annotated_image=annotated_rel_path,
        image_path=latest_image_filename,
        yolo_boxes=yolo_boxes,
        yolo_available=YOLO_AVAILABLE,
        yolo_warning=yolo_warning,
        yolo_weights=YOLO_WEIGHTS_PATH,
    )


@app.route('/verification_history')
def verification_history():
    """View verification history"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get daily history
    daily_history = get_daily_history()
    
    # Get all verification images
    verification_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'verification')
    images = []
    if os.path.exists(verification_dir):
        for filename in sorted(os.listdir(verification_dir), reverse=True)[:50]:  # Last 50
            if allowed_file(filename):
                images.append({
                    'filename': filename,
                    'path': os.path.join('verification', filename),
                    'time': datetime.fromtimestamp(
                        os.path.getmtime(os.path.join(verification_dir, filename))
                    ).strftime('%Y-%m-%d %H:%M:%S')
                })
    
    return render_template('verification_history.html', 
                         images=images,
                         daily_history=daily_history)


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Serve uploaded images"""
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename))


@app.route('/test_email')
def test_email():
    """Test email configuration"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_email = session.get('user_email')
    
    # Check configuration
    config_status = {
        'mail_server': MAIL_SERVER,
        'mail_port': MAIL_PORT,
        'mail_username': MAIL_USERNAME[:3] + '***' if len(MAIL_USERNAME) > 3 else 'NOT SET',
        'mail_password_set': 'YES' if MAIL_PASSWORD != 'your-app-password' else 'NO',
    }
    
    # Try to send test email
    test_image = os.path.join(app.config['UPLOAD_FOLDER'], 'verification')
    if os.path.exists(test_image) and os.listdir(test_image):
        # Use first available image
        test_img = os.path.join(test_image, os.listdir(test_image)[0])
    else:
        test_img = None
    
    if test_img:
        result, message = send_email_alert(user_email, 'TEST123', test_img, is_registered=False)
    else:
        # Create a simple test without image
        try:
            msg = MIMEMultipart()
            msg['From'] = MAIL_USERNAME
            msg['To'] = user_email
            msg['Subject'] = "Test Email - Vehicle Recognition System"
            body = "This is a test email from the Vehicle License Plate Recognition System. If you receive this, your email configuration is working!"
            msg.attach(MIMEText(body, 'plain'))
            
            server = smtplib.SMTP(MAIL_SERVER, MAIL_PORT)
            server.starttls()
            server.login(MAIL_USERNAME, MAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            result = True
            message = "Test email sent successfully!"
        except Exception as e:
            result = False
            message = f"Error: {str(e)}"
   
    
    return render_template('test_email.html', 
                         config_status=config_status,
                         test_result=result,
                         test_message=message,
                         user_email=user_email)


@app.route('/api/parking-slots', methods=['GET'])
def get_parking_slots():
    """API endpoint for real-time parking slot data"""
    import json
    import requests
    
    try:
        # Firebase Realtime Database URL - fetch parking_status
        firebase_url = "https://lite-bc5e2-default-rtdb.asia-southeast1.firebasedatabase.app/parking_status.json"
        
        # Fetch data from Firebase
        response = requests.get(firebase_url, timeout=5)
        
        if response.status_code == 200:
            parking_status = response.json()
            
            if parking_status and isinstance(parking_status, dict):
                # Extract data from parking_status
                total_spaces = parking_status.get('total_spaces', 2)
                occupied_spaces = parking_status.get('occupied_spaces', 0)
                empty_spaces = parking_status.get('empty_spaces', 2)
                status = parking_status.get('status', 'EMPTY')
                timestamp = parking_status.get('timestamp', '')
                occupancy_rate = parking_status.get('occupancy_rate', 0)
                
                # Create 2 parking slots based on occupied_spaces
                slots = []
                for i in range(1, total_spaces + 1):
                    if i <= occupied_spaces:
                        slot_status = 'occupied'
                    else:
                        slot_status = 'available'
                    
                    slots.append({
                        'slot_number': f'Slot {i}',
                        'status': slot_status,
                        'vehicle_number': '',
                        'timestamp': timestamp
                    })
                
                return json.dumps({
                    'success': True,
                    'total_slots': total_spaces,
                    'available_slots': empty_spaces,
                    'occupied_slots': occupied_spaces,
                    'occupancy_rate': occupancy_rate,
                    'status': status,
                    'slots': slots,
                    'timestamp': datetime.now().isoformat()
                }), 200, {'Content-Type': 'application/json'}
            else:
                # No data in Firebase
                return json.dumps({
                    'success': False,
                    'message': 'No parking_status data found in Firebase',
                    'total_slots': 0,
                    'available_slots': 0,
                    'occupied_slots': 0,
                    'slots': []
                }), 200, {'Content-Type': 'application/json'}
        else:
            # Error fetching from Firebase
            return json.dumps({
                'success': False,
                'message': f'Firebase connection failed: {response.status_code}',
                'total_slots': 0,
                'available_slots': 0,
                'occupied_slots': 0,
                'slots': []
            }), 200, {'Content-Type': 'application/json'}
            
    except Exception as e:
        print(f"Error fetching from Firebase: {e}")
        return json.dumps({
            'success': False,
            'message': f'Error: {str(e)}',
            'total_slots': 0,
            'available_slots': 0,
            'occupied_slots': 0,
            'slots': []
        }), 200, {'Content-Type': 'application/json'}


if __name__ == '__main__':
    print("=" * 60)
    print("Vehicle License Plate Recognition Web Application")
    print("=" * 60)
    print(f"Starting server on http://127.0.0.1:5000")
    print(f"API Key: {DEFAULT_API_KEY[:20]}...")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)

